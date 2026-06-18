from __future__ import annotations

from io import BytesIO
from typing import cast

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

CONSTRAINT_HEADERS = ["", "値", "単位"]
FACILITY_CONDITION_HEADERS = [
    "施設コード",
    "施設名",
    "CPA",
    "除外対象フラグ",
    "月ごとの開催上限",
    "実施可能曜日",
]
FACILITY_DAILY_TARGET_HEADERS = [
    "施設コード",
    "施設名",
    "POレベル",
    "支社名",
    "支店名",
    "日付",
    "日付フラグ",
    "標準目標値(季節加味)",
]
DATE_MASTER_HEADERS = ["日付", "曜日（週番号）", "日付種別"]


def split_facility_targets(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    facility_summary_df = df[
        ["FACILITY_CODE", "FACILITY_NAME", "CPA", "IS_EXCLUDED"]
    ].drop_duplicates(subset=["FACILITY_CODE"])

    remaining_columns = [col for col in df.columns if col not in ["CPA", "IS_EXCLUDED"]]
    facility_details_df = df[remaining_columns]

    return facility_summary_df, facility_details_df


def apply_basic_table_style(
    *,
    ws: Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    """Excel表に最低限の罫線・配置を適用する。色は使わない。"""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for cell in ws[min_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_headers(ws: Worksheet, headers: list[str], *, row_idx: int = 1) -> None:
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row_idx, column=col_idx, value=header)


def write_dataframe_rows(
    ws: Worksheet, df: pd.DataFrame, *, start_row: int = 2
) -> None:
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def set_row_heights(
    ws: Worksheet, *, min_row: int, max_row: int, height: float
) -> None:
    for row_idx in range(min_row, max_row + 1):
        ws.row_dimensions[row_idx].height = height


def set_column_widths(ws: Worksheet, widths: dict[str, float]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def create_constraint_sheet(
    *,
    wb: Workbook,
    proposal_period: str,
    monthly_event_count: int | None,
    weekday_pattern: str,
    target_pi: int | None,
    condition_cost: int | None,
) -> None:
    ws = cast(Worksheet, wb.active)
    ws.title = "制約条件"

    rows = [
        ("提案期間", proposal_period, ""),
        ("月辺り実施回数（稼働ライン）", monthly_event_count, "回"),
        ("稼働曜日", weekday_pattern, ""),
        ("目標実績", target_pi, "PI"),
        ("条件コスト", condition_cost, "円"),
        ("目標CPA水準", None, "円"),
    ]

    write_headers(ws, CONSTRAINT_HEADERS)

    for row_idx, (label, value, unit) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=3, value=unit)

    ws["B7"] = '=IFERROR(ROUND(B6/B5,0),"")'

    # ----- スタイル付与 -----

    apply_basic_table_style(ws=ws, min_row=1, max_row=7, min_col=1, max_col=3)

    for row_idx in range(2, 8):
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        ws.cell(row=row_idx, column=2).alignment = Alignment(
            horizontal="left",
            vertical="center",
        )
        ws.cell(row=row_idx, column=3).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    set_column_widths(ws, {"A": 32, "B": 24, "C": 8})
    set_row_heights(ws, min_row=1, max_row=7, height=24)

    for cell_address in ["B3", "B5", "B6", "B7"]:
        ws[cell_address].number_format = "#,##0"


def create_facility_condition_sheet(
    *,
    wb: Workbook,
    facility_conditions_df: pd.DataFrame,
) -> None:
    ws = wb.create_sheet(title="施設別制約条件")

    write_headers(ws, FACILITY_CONDITION_HEADERS)
    write_dataframe_rows(ws, facility_conditions_df)

    # ----- スタイル付与 -----

    max_row = max(len(facility_conditions_df) + 1, 2)
    max_col = len(FACILITY_CONDITION_HEADERS)

    apply_basic_table_style(
        ws=ws,
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    )

    set_column_widths(
        ws,
        {
            "A": 16,
            "B": 40,
            "C": 16,
            "D": 16,
            "E": 16,
            "F": 16,
        },
    )
    set_row_heights(ws, min_row=1, max_row=max_row, height=24)

    for row_idx in range(2, max_row + 1):
        ws.cell(row=row_idx, column=3).number_format = "#,##0"


def create_facility_daily_target_sheet(
    *,
    wb: Workbook,
    facility_targets_df: pd.DataFrame,
) -> None:
    ws = wb.create_sheet(title="施設別日別目標値")

    write_headers(ws, FACILITY_DAILY_TARGET_HEADERS)
    write_dataframe_rows(ws, facility_targets_df)

    # ----- スタイル付与 -----

    max_row = max(len(facility_targets_df) + 1, 2)
    max_col = len(FACILITY_DAILY_TARGET_HEADERS)

    apply_basic_table_style(
        ws=ws,
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    )

    set_column_widths(
        ws,
        {
            "A": 16,
            "B": 40,
            "C": 16,
            "D": 16,
            "E": 16,
            "F": 16,
            "G": 16,
            "H": 24,
        },
    )
    set_row_heights(ws, min_row=1, max_row=max_row, height=24)

    for row_idx in range(2, max_row + 1):
        ws.cell(row=row_idx, column=8).number_format = "#,##0"


def create_date_master_sheet(*, wb: Workbook, date_master_df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title="日付情報")

    write_headers(ws, DATE_MASTER_HEADERS)
    write_dataframe_rows(ws, date_master_df)

    # ----- スタイル付与 -----

    max_row = max(len(date_master_df) + 1, 2)
    max_col = len(DATE_MASTER_HEADERS)

    apply_basic_table_style(
        ws=ws,
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    )

    set_column_widths(ws, {"A": 16, "B": 16, "C": 16})
    set_row_heights(ws, min_row=1, max_row=max_row, height=24)


def create_output_format_sheet(*, wb: Workbook, date_master_df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title="アウトプットデータ形式")

    fixed_headers = [
        "施設名",
        "POレベル",
        "支社",
        "支店",
        "除外対象フラグ",
    ]

    fixed_start_col = 2
    date_start_col = fixed_start_col + len(fixed_headers)

    data_start_row = 6
    data_end_row = 18

    max_row = data_end_row
    max_col = date_start_col + len(date_master_df) - 1

    ws.cell(row=2, column=date_start_col, value="日付")
    ws.cell(row=2, column=date_start_col + 1, value="日付種別")

    for col_idx, header in enumerate(fixed_headers, start=fixed_start_col):
        ws.cell(row=5, column=col_idx, value=header)

    for offset, row in enumerate(date_master_df.itertuples(index=False)):
        col_idx = date_start_col + offset

        date_value = pd.to_datetime(row.DATE).date()

        date_cell = ws.cell(row=3, column=col_idx, value=date_value)
        date_cell.number_format = "yyyy/m/d"

        ws.cell(
            row=4,
            column=col_idx,
            value=row.WEEKDAY_NAME_AND_WEEK_NUMBER_MONTHLY,
        )

        ws.cell(
            row=5,
            column=col_idx,
            value=row.DATE_FLAG,
        )

    apply_basic_table_style(
        ws=ws,
        min_row=2,
        max_row=max_row,
        min_col=fixed_start_col,
        max_col=max_col,
    )

    for col_idx in range(fixed_start_col, date_start_col):
        ws.cell(row=5, column=col_idx).font = Font(bold=True)

    for row_idx in range(data_start_row, data_end_row + 1):
        for col_idx in range(fixed_start_col, date_start_col):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="left",
                vertical="center",
            )

    for col_idx in range(fixed_start_col, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    set_row_heights(ws, min_row=1, max_row=max_row, height=16)


def build_input_excel(
    *,
    proposal_period: str,
    monthly_event_count: int | None,
    weekday_pattern: str,
    target_pi: int | None,
    condition_cost: int | None,
    date_master_df: pd.DataFrame,
    facility_targets_raw_df: pd.DataFrame,
) -> bytes:
    wb = Workbook()
    facility_conditions_df, facility_targets_df = split_facility_targets(
        facility_targets_raw_df
    )

    create_constraint_sheet(
        wb=wb,
        proposal_period=proposal_period,
        monthly_event_count=monthly_event_count,
        weekday_pattern=weekday_pattern,
        target_pi=target_pi,
        condition_cost=condition_cost,
    )
    create_facility_condition_sheet(
        wb=wb,
        facility_conditions_df=facility_conditions_df,
    )
    create_facility_daily_target_sheet(wb=wb, facility_targets_df=facility_targets_df)
    create_date_master_sheet(wb=wb, date_master_df=date_master_df)
    create_output_format_sheet(wb=wb, date_master_df=date_master_df)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
