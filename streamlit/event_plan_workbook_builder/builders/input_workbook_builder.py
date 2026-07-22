from __future__ import annotations
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from entities import (
    ConstraintDetail,
    DateDetail,
    FacilityDailyTargetDetail,
    FacilityDetail,
)
from config import COPILOT_INPUT_TEMPLATE_PATH
from utils import calculate_input_data_cpa


class InputWorkbookBuilder:
    COMMON_CONSTRAINT_SHEET_NAME = "共通制約条件"

    REGIONAL_OFFICE_CONSTRAINT_SHEET_NAME = "支社別制約条件"
    REGIONAL_OFFICE_OPERATING_DAYS_START_ROW = 4
    REGIONAL_OFFICE_DAILY_EVENT_LIMITS_START_ROW = 4

    FACILITY_CONDITION_SHEET_NAME = "施設別制約条件"
    FACILITY_CONDITION_START_ROW = 3

    FACILITY_DAILY_TARGET_SHEET_NAME = "施設別・日別_目標値"
    FACILITY_DAILY_TARGET_START_ROW = 3

    DATE_SHEET_NAME = "日付情報"
    DATE_START_ROW = 3

    OUTPUT_SHEET_NAME = "アウトプットデータ形式"
    OUTPUT_START_COL = 3

    def __init__(
        self,
        *,
        constraint_details: list[ConstraintDetail],
        facility_details: list[FacilityDetail],
        facility_daily_target_details: list[FacilityDailyTargetDetail],
        date_details: list[DateDetail],
        is_all_regional_offices: bool = False,
    ) -> None:
        wb = load_workbook(COPILOT_INPUT_TEMPLATE_PATH)
        template_buffer = BytesIO()
        wb.save(template_buffer)

        self.template_workbook_bytes = template_buffer.getvalue()
        self.wb = wb
        self.constraint_details = constraint_details
        self.facility_details = facility_details
        self.facility_daily_target_details = facility_daily_target_details
        self.date_details = date_details
        self.input_data_cpa = calculate_input_data_cpa(facility_details)
        self.is_all_regional_offices = is_all_regional_offices

    def _write_common_constraint_sheet(
        self, ws: Worksheet, constraint_detail: ConstraintDetail
    ) -> None:

        ws[f"C3"] = constraint_detail.proposal_period
        ws[f"C4"] = constraint_detail.daily_event_limit
        ws[f"C5"] = constraint_detail.weekday_pattern
        ws[f"C6"] = constraint_detail.target_actual
        ws[f"C7"] = constraint_detail.constraint_cost
        ws[f"C8"] = constraint_detail.target_cpa()
        ws[f"C9"] = self.input_data_cpa

    def _write_regional_office_constraint_sheet(self, ws: Worksheet) -> None:
        for i, constraint in enumerate(self.constraint_details):
            weekday_pattern_row = self.REGIONAL_OFFICE_OPERATING_DAYS_START_ROW + i
            daily_event_limits_row = (
                self.REGIONAL_OFFICE_DAILY_EVENT_LIMITS_START_ROW + i
            )

            ws[f"B{weekday_pattern_row}"] = constraint.regional_office
            ws[f"C{weekday_pattern_row}"] = constraint.weekday_pattern

            ws[f"E{daily_event_limits_row}"] = constraint.regional_office
            ws[f"F{daily_event_limits_row}"] = constraint.daily_event_limit - 5
            ws[f"G{daily_event_limits_row}"] = constraint.daily_event_limit

    def _write_facility_condition_sheet(
        self, ws: Worksheet, facility_details: list[FacilityDetail]
    ) -> None:
        for (
            row_offset,
            facility_detail,
        ) in enumerate(facility_details):
            row_idx = self.FACILITY_CONDITION_START_ROW + row_offset

            ws[f"B{row_idx}"] = facility_detail.facility_code
            ws[f"C{row_idx}"] = facility_detail.facility_name
            ws[f"D{row_idx}"] = facility_detail.cpa
            ws[f"E{row_idx}"] = facility_detail.monthly_event_limit
            ws[f"F{row_idx}"] = facility_detail.operating_days

    def _write_facility_daily_target_sheet(
        self,
        ws: Worksheet,
        facility_daily_target_details: list[FacilityDailyTargetDetail],
    ) -> None:
        for row_offset, facility_daily_target_detail in enumerate(
            facility_daily_target_details
        ):
            row_idx = self.FACILITY_DAILY_TARGET_START_ROW + row_offset

            ws[f"B{row_idx}"] = facility_daily_target_detail.facility_code
            ws[f"C{row_idx}"] = facility_daily_target_detail.facility_name
            ws[f"D{row_idx}"] = facility_daily_target_detail.po_level
            ws[f"E{row_idx}"] = facility_daily_target_detail.regional_office
            ws[f"F{row_idx}"] = facility_daily_target_detail.branch_office
            ws[f"G{row_idx}"] = facility_daily_target_detail.date
            ws[f"H{row_idx}"] = facility_daily_target_detail.date_flag
            ws[f"I{row_idx}"] = facility_daily_target_detail.target_value

    def _write_date_sheet(self, ws: Worksheet) -> None:
        for row_offset, date_detail in enumerate(self.date_details):
            row_idx = self.DATE_START_ROW + row_offset

            ws[f"B{row_idx}"] = date_detail.date
            ws[f"C{row_idx}"] = date_detail.weekday_name_and_week_number_monthly
            ws[f"D{row_idx}"] = date_detail.date_flag

    def _write_output_template_sheet_date_header(self, ws: Worksheet) -> None:
        for col_offset, date_detail in enumerate(self.date_details):
            col_idx = self.OUTPUT_START_COL + col_offset

            ws.cell(row=3, column=col_idx, value=date_detail.date)
            ws.cell(
                row=4,
                column=col_idx,
                value=date_detail.weekday_name_and_week_number_monthly,
            )
            ws.cell(row=5, column=col_idx, value=date_detail.date_flag)

    def build(self) -> bytes:
        regional_facility_map: dict[str, list[FacilityDetail]] = {}

        for facility_detail in self.facility_details:
            key = facility_detail.regional_office

            if key not in regional_facility_map:
                regional_facility_map[key] = []

            regional_facility_map[key].append(facility_detail)

        regional_facility_daily_target_detail_map: dict[
            str, list[FacilityDailyTargetDetail]
        ] = {}

        for facility_daily_target_detail in self.facility_daily_target_details:
            key = facility_daily_target_detail.regional_office

            if key not in regional_facility_daily_target_detail_map:
                regional_facility_daily_target_detail_map[key] = []

            regional_facility_daily_target_detail_map[key].append(
                facility_daily_target_detail
            )

        def add_workbook_to_zip(
            wb: Workbook,
            zip_file: ZipFile,
            *,
            constraint_detail: ConstraintDetail,
            is_regional_version: bool = True,
        ) -> None:
            regional_office = constraint_detail.regional_office

            facility_details = (
                regional_facility_map[regional_office]
                if is_regional_version
                else self.facility_details
            )

            daily_target_details = (
                regional_facility_daily_target_detail_map[regional_office]
                if is_regional_version
                else self.facility_daily_target_details
            )

            output_filename = (
                f"{regional_office}.xlsx" if is_regional_version else "_全国版.xlsx"
            )

            try:
                self._write_common_constraint_sheet(
                    wb[self.COMMON_CONSTRAINT_SHEET_NAME],
                    constraint_detail,
                )

                if is_regional_version:
                    regional_office_constraint_ws = wb[
                        self.REGIONAL_OFFICE_CONSTRAINT_SHEET_NAME
                    ]
                    wb.remove(regional_office_constraint_ws)
                else:
                    self._write_regional_office_constraint_sheet(
                        wb[self.REGIONAL_OFFICE_CONSTRAINT_SHEET_NAME]
                    )

                self._write_facility_condition_sheet(
                    wb[self.FACILITY_CONDITION_SHEET_NAME],
                    facility_details,
                )
                self._write_facility_daily_target_sheet(
                    wb[self.FACILITY_DAILY_TARGET_SHEET_NAME],
                    daily_target_details,
                )
                self._write_date_sheet(wb[self.DATE_SHEET_NAME])
                self._write_output_template_sheet_date_header(
                    wb[self.OUTPUT_SHEET_NAME]
                )

                with BytesIO() as workbook_buffer:
                    wb.save(workbook_buffer)
                    zip_file.writestr(
                        output_filename,
                        workbook_buffer.getvalue(),
                    )
            finally:
                wb.close()

        zip_buffer = BytesIO()
        with ZipFile(
            zip_buffer,
            mode="w",
            compression=ZIP_DEFLATED,
        ) as zip_file:
            total_daily_event_limit = 0
            total_target_actual = 0
            total_constraint_cost = 0

            # 支社別版のExcelを作成する
            for constraint_detail in self.constraint_details:
                wb = load_workbook(BytesIO(self.template_workbook_bytes))
                add_workbook_to_zip(wb, zip_file, constraint_detail=constraint_detail)

                total_daily_event_limit += constraint_detail.daily_event_limit
                total_target_actual += constraint_detail.target_actual
                total_constraint_cost += constraint_detail.constraint_cost

            # 全支社版のExcelを作成
            wb = load_workbook(BytesIO(self.template_workbook_bytes))
            add_workbook_to_zip(
                wb,
                zip_file,
                constraint_detail=ConstraintDetail(
                    regional_office="",
                    proposal_period=self.constraint_details[0].proposal_period,
                    daily_event_limit=total_daily_event_limit,
                    weekday_pattern="",
                    target_actual=total_target_actual,
                    constraint_cost=total_constraint_cost,
                ),
                is_regional_version=False,
            )

        return zip_buffer.getvalue()
