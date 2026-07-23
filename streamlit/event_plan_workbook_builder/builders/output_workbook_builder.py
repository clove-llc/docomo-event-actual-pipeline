from __future__ import annotations
from io import BytesIO
from typing import Iterable
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from entities import (
    ConstraintDetail,
    DateDetail,
    FacilityDailyTargetDetail,
    FacilityDetail,
)
from config import COPILOT_OUTPUT_TEMPLATE_PATH


def build_rank_eq_map(
    values: Iterable[int | None],
    *,
    reverse: bool = True,
) -> dict[int, int]:
    """ExcelのRANK.EQ相当の順位マップを作る。

    reverse=True  : 大きい値ほど1位
    reverse=False : 小さい値ほど1位
    """
    non_null_values = [value for value in values if value is not None]

    sorted_values = sorted(non_null_values, reverse=reverse)

    rank_map: dict[int, int] = {}

    for index, value in enumerate(sorted_values):
        if value not in rank_map:
            rank_map[value] = index + 1

    return rank_map


def get_rank(value: int | None, rank_map: dict[int, int]) -> int | None:
    if value is None:
        return None

    return rank_map[value]


class OutputWorkbookBuilder:
    COMMON_CONSTRAINT_SHEET_NAME = "共通制約条件"

    REGIONAL_OFFICE_CONSTRAINT_SHEET_NAME = "支社別制約条件"
    REGIONAL_OFFICE_OPERATING_DAYS_START_ROW = 4
    REGIONAL_OFFICE_DAILY_EVENT_LIMITS_START_ROW = 4

    COPILOT_OUTPUT_SHEET_NAME = "Copilotの出力結果"

    FACILITY_OUTPUT_SHEET_NAMES = [
        "想定実績値_集計",
        "想定コスト_集計",
        "想定CPA_集計",
        "最適なスタッフ数_集計",
    ]
    FACILITY_OUTPUT_START_ROW = 8

    FACILITY_DAILY_TARGET_SHEET_NAME = "施設別・日別_目標値"
    FACILITY_DAILY_TARGET_START_ROW = 3

    FACILITY_PRIORITY_SHEET_NAME = "優先度_ベース"
    FACILITY_PRIORITY_START_ROW = 3

    def __init__(
        self,
        *,
        constraint_details: list[ConstraintDetail],
        facility_details: list[FacilityDetail],
        cpa_avg: int,
        date_details: list[DateDetail],
        facility_daily_target_details: list[FacilityDailyTargetDetail],
    ) -> None:
        self.template_workbook_bytes = COPILOT_OUTPUT_TEMPLATE_PATH.read_bytes()
        self.constraint_details = constraint_details
        self.facility_details = facility_details
        self.cpa_avg = cpa_avg
        self.date_details = date_details
        self.facility_daily_target_details = facility_daily_target_details

    def _write_common_constraint_sheet(
        self, ws: Worksheet, constraint_detail: ConstraintDetail
    ) -> None:
        ws[f"C3"] = constraint_detail.proposal_period
        ws[f"C4"] = constraint_detail.daily_event_limit
        ws[f"C5"] = constraint_detail.weekday_pattern
        ws[f"C6"] = constraint_detail.target_actual
        ws[f"C7"] = constraint_detail.constraint_cost
        ws[f"C8"] = constraint_detail.target_cpa()
        ws[f"C9"] = self.cpa_avg

    def _write_regional_office_constraint_sheet(self, ws: Worksheet) -> None:
        for i, constraint in enumerate(self.constraint_details):
            weekday_pattern_row = self.REGIONAL_OFFICE_OPERATING_DAYS_START_ROW + i
            daily_event_limits_row = (
                self.REGIONAL_OFFICE_DAILY_EVENT_LIMITS_START_ROW + i
            )

            ws[f"B{weekday_pattern_row}"] = constraint.regional_office
            ws[f"C{weekday_pattern_row}"] = constraint.weekday_pattern

            ws[f"E{daily_event_limits_row}"] = constraint.regional_office
            ws[f"F{daily_event_limits_row}"] = constraint.daily_event_limit - 5
            ws[f"G{daily_event_limits_row}"] = constraint.daily_event_limit

    def _write_facility_output_sheet(
        self, ws: Worksheet, facility_details: list[FacilityDetail]
    ) -> None:
        for (
            row_offset,
            facility_detail,
        ) in enumerate(facility_details):
            row_idx = self.FACILITY_OUTPUT_START_ROW + row_offset

            ws[f"B{row_idx}"] = facility_detail.facility_name
            ws[f"C{row_idx}"] = facility_detail.po_level
            ws[f"D{row_idx}"] = facility_detail.regional_office
            ws[f"E{row_idx}"] = facility_detail.branch_office
            ws[f"F{row_idx}"] = facility_detail.cpa
            ws[f"G{row_idx}"] = facility_detail.monthly_event_limit
            ws[f"H{row_idx}"] = facility_detail.operating_days

    def _write_facility_daily_target_sheet(
        self,
        ws: Worksheet,
        facility_daily_target_details: list[FacilityDailyTargetDetail],
    ) -> None:
        for row_offset, facility_daily_target_detail in enumerate(
            facility_daily_target_details
        ):
            row_idx = self.FACILITY_DAILY_TARGET_START_ROW + row_offset

            ws[f"B{row_idx}"] = facility_daily_target_detail.search_key()
            ws[f"C{row_idx}"] = facility_daily_target_detail.facility_name
            ws[f"D{row_idx}"] = facility_daily_target_detail.date
            ws[f"E{row_idx}"] = facility_daily_target_detail.date_flag
            ws[f"F{row_idx}"] = facility_daily_target_detail.target_value

    def _write_date_header(self, ws: Worksheet, start_col: int = 9) -> None:
        for col_offset, date_detail in enumerate(self.date_details):
            col_idx = start_col + col_offset

            ws.cell(row=5, column=col_idx, value=date_detail.date)
            ws.cell(
                row=6,
                column=col_idx,
                value=date_detail.weekday_name_and_week_number_monthly,
            )
            ws.cell(row=7, column=col_idx, value=date_detail.date_flag)

    def _write_facility_priority_sheet(
        self, ws: Worksheet, facility_details: list[FacilityDetail]
    ) -> None:
        # ----- CPAランクマップを作成（小さいほどランクが高い） -----
        cpa_rank_map = build_rank_eq_map(
            (detail.cpa for detail in facility_details),
            reverse=False,
        )

        # ----- 日付フラグ別目標値ランクマップを作成（大きいほどランクが高い） -----
        weekday_rank_map = build_rank_eq_map(
            detail.avg_weekday_standard_target_seasonal for detail in facility_details
        )
        regular_weekend_rank_map = build_rank_eq_map(
            detail.avg_regular_weekend_standard_target_seasonal
            for detail in facility_details
        )
        three_day_holiday_rank_map = build_rank_eq_map(
            detail.avg_three_day_holiday_standard_target_seasonal
            for detail in facility_details
        )
        bridge_holiday_rank_map = build_rank_eq_map(
            detail.avg_bridge_holiday_standard_target_seasonal
            for detail in facility_details
        )
        gw_rank_map = build_rank_eq_map(
            detail.avg_gw_standard_target_seasonal for detail in facility_details
        )
        obon_rank_map = build_rank_eq_map(
            detail.avg_obon_standard_target_seasonal for detail in facility_details
        )
        new_year_rank_map = build_rank_eq_map(
            detail.avg_new_year_standard_target_seasonal for detail in facility_details
        )
        year_end_rank_map = build_rank_eq_map(
            detail.avg_year_end_standard_target_seasonal for detail in facility_details
        )
        black_friday_rank_map = build_rank_eq_map(
            detail.avg_black_friday_standard_target_seasonal
            for detail in facility_details
        )

        # ----- 1周目: 各施設の日付フラグ別目標値ランク・（CPAランク + 日付フラグ別）のランクを作成 -----
        facility_rank_rows = []

        for facility_detail in facility_details:
            cpa_rank = get_rank(facility_detail.cpa, cpa_rank_map)

            weekday_rank = get_rank(
                facility_detail.avg_weekday_standard_target_seasonal,
                weekday_rank_map,
            )
            regular_weekend_rank = get_rank(
                facility_detail.avg_regular_weekend_standard_target_seasonal,
                regular_weekend_rank_map,
            )
            three_day_holiday_rank = get_rank(
                facility_detail.avg_three_day_holiday_standard_target_seasonal,
                three_day_holiday_rank_map,
            )
            bridge_holiday_rank = get_rank(
                facility_detail.avg_bridge_holiday_standard_target_seasonal,
                bridge_holiday_rank_map,
            )
            gw_rank = get_rank(
                facility_detail.avg_gw_standard_target_seasonal,
                gw_rank_map,
            )
            obon_rank = get_rank(
                facility_detail.avg_obon_standard_target_seasonal,
                obon_rank_map,
            )
            new_year_rank = get_rank(
                facility_detail.avg_new_year_standard_target_seasonal,
                new_year_rank_map,
            )
            year_end_rank = get_rank(
                facility_detail.avg_year_end_standard_target_seasonal,
                year_end_rank_map,
            )
            black_friday_rank = get_rank(
                facility_detail.avg_black_friday_standard_target_seasonal,
                black_friday_rank_map,
            )

            def sum_rank(
                rank_1: int | None,
                rank_2: int | None,
            ) -> int | None:
                if rank_1 is None or rank_2 is None:
                    return None

                return rank_1 + rank_2

            weekday_sum_rank = sum_rank(cpa_rank, weekday_rank)
            regular_weekend_sum_rank = sum_rank(cpa_rank, regular_weekend_rank)
            three_day_holiday_sum_rank = sum_rank(cpa_rank, three_day_holiday_rank)
            bridge_holiday_sum_rank = sum_rank(cpa_rank, bridge_holiday_rank)
            gw_sum_rank = sum_rank(cpa_rank, gw_rank)
            obon_sum_rank = sum_rank(cpa_rank, obon_rank)
            new_year_sum_rank = sum_rank(cpa_rank, new_year_rank)
            year_end_sum_rank = sum_rank(cpa_rank, year_end_rank)
            black_friday_sum_rank = sum_rank(cpa_rank, black_friday_rank)

            facility_rank_rows.append(
                {
                    "facility_detail": facility_detail,
                    "cpa_rank": cpa_rank,
                    "weekday_rank": weekday_rank,
                    "regular_weekend_rank": regular_weekend_rank,
                    "three_day_holiday_rank": three_day_holiday_rank,
                    "bridge_holiday_rank": bridge_holiday_rank,
                    "gw_rank": gw_rank,
                    "obon_rank": obon_rank,
                    "new_year_rank": new_year_rank,
                    "year_end_rank": year_end_rank,
                    "black_friday_rank": black_friday_rank,
                    "weekday_sum_rank": weekday_sum_rank,
                    "regular_weekend_sum_rank": regular_weekend_sum_rank,
                    "three_day_holiday_sum_rank": three_day_holiday_sum_rank,
                    "bridge_holiday_sum_rank": bridge_holiday_sum_rank,
                    "gw_sum_rank": gw_sum_rank,
                    "obon_sum_rank": obon_sum_rank,
                    "new_year_sum_rank": new_year_sum_rank,
                    "year_end_sum_rank": year_end_sum_rank,
                    "black_friday_sum_rank": black_friday_sum_rank,
                }
            )

        # ----- sum_rankのランクマップを作成（小さいほど高い） -----
        weekday_priority_rank_map = build_rank_eq_map(
            (row["weekday_sum_rank"] for row in facility_rank_rows),
            reverse=False,
        )
        regular_weekend_priority_rank_map = build_rank_eq_map(
            (row["regular_weekend_sum_rank"] for row in facility_rank_rows),
            reverse=False,
        )
        three_day_holiday_priority_rank_map = build_rank_eq_map(
            (row["three_day_holiday_sum_rank"] for row in facility_rank_rows),
            reverse=False,
        )
        bridge_holiday_priority_rank_map = build_rank_eq_map(
            (row["bridge_holiday_sum_rank"] for row in facility_rank_rows),
            reverse=False,
        )
        gw_priority_rank_map = build_rank_eq_map(
            (row["gw_sum_rank"] for row in facility_rank_rows),
            reverse=False,
        )
        obon_priority_rank_map = build_rank_eq_map(
            (row["obon_sum_rank"] for row in facility_rank_rows),
            reverse=False,
        )
        new_year_priority_rank_map = build_rank_eq_map(
            (row["new_year_sum_rank"] for row in facility_rank_rows),
            reverse=False,
        )
        year_end_priority_rank_map = build_rank_eq_map(
            (row["year_end_sum_rank"] for row in facility_rank_rows),
            reverse=False,
        )
        black_friday_priority_rank_map = build_rank_eq_map(
            (row["black_friday_sum_rank"] for row in facility_rank_rows),
            reverse=False,
        )

        # ----- 2周目: Excelへ書き込む -----
        for row_offset, rank_row in enumerate(facility_rank_rows):
            row_idx = self.FACILITY_PRIORITY_START_ROW + row_offset
            facility_detail = rank_row["facility_detail"]

            # 施設情報
            ws[f"B{row_idx}"] = facility_detail.facility_name
            ws[f"C{row_idx}"] = facility_detail.po_level
            ws[f"D{row_idx}"] = facility_detail.regional_office
            ws[f"E{row_idx}"] = facility_detail.branch_office
            ws[f"F{row_idx}"] = facility_detail.cpa
            ws[f"G{row_idx}"] = facility_detail.monthly_event_limit
            ws[f"H{row_idx}"] = facility_detail.operating_days

            # 日付フラグ別の目標値
            ws[f"I{row_idx}"] = facility_detail.avg_weekday_standard_target_seasonal
            ws[f"J{row_idx}"] = (
                facility_detail.avg_regular_weekend_standard_target_seasonal
            )
            ws[f"K{row_idx}"] = (
                facility_detail.avg_three_day_holiday_standard_target_seasonal
            )
            ws[f"L{row_idx}"] = (
                facility_detail.avg_bridge_holiday_standard_target_seasonal
            )
            ws[f"M{row_idx}"] = facility_detail.avg_gw_standard_target_seasonal
            ws[f"N{row_idx}"] = facility_detail.avg_obon_standard_target_seasonal
            ws[f"O{row_idx}"] = facility_detail.avg_new_year_standard_target_seasonal
            ws[f"P{row_idx}"] = facility_detail.avg_year_end_standard_target_seasonal
            ws[f"Q{row_idx}"] = (
                facility_detail.avg_black_friday_standard_target_seasonal
            )

            # CPAランク
            ws[f"R{row_idx}"] = rank_row["cpa_rank"]

            # 日付フラグ別ランク
            ws[f"S{row_idx}"] = rank_row["weekday_rank"]
            ws[f"T{row_idx}"] = rank_row["regular_weekend_rank"]
            ws[f"U{row_idx}"] = rank_row["three_day_holiday_rank"]
            ws[f"V{row_idx}"] = rank_row["bridge_holiday_rank"]
            ws[f"W{row_idx}"] = rank_row["gw_rank"]
            ws[f"X{row_idx}"] = rank_row["obon_rank"]
            ws[f"Y{row_idx}"] = rank_row["new_year_rank"]
            ws[f"Z{row_idx}"] = rank_row["year_end_rank"]
            ws[f"AA{row_idx}"] = rank_row["black_friday_rank"]

            # CPAランク + 日付フラグ別ランク
            ws[f"AB{row_idx}"] = rank_row["weekday_sum_rank"]
            ws[f"AC{row_idx}"] = rank_row["regular_weekend_sum_rank"]
            ws[f"AD{row_idx}"] = rank_row["three_day_holiday_sum_rank"]
            ws[f"AE{row_idx}"] = rank_row["bridge_holiday_sum_rank"]
            ws[f"AF{row_idx}"] = rank_row["gw_sum_rank"]
            ws[f"AG{row_idx}"] = rank_row["obon_sum_rank"]
            ws[f"AH{row_idx}"] = rank_row["new_year_sum_rank"]
            ws[f"AI{row_idx}"] = rank_row["year_end_sum_rank"]
            ws[f"AJ{row_idx}"] = rank_row["black_friday_sum_rank"]

            # sum_rankの最終順位
            ws[f"AK{row_idx}"] = get_rank(
                rank_row["weekday_sum_rank"],
                weekday_priority_rank_map,
            )
            ws[f"AL{row_idx}"] = get_rank(
                rank_row["regular_weekend_sum_rank"],
                regular_weekend_priority_rank_map,
            )
            ws[f"AM{row_idx}"] = get_rank(
                rank_row["three_day_holiday_sum_rank"],
                three_day_holiday_priority_rank_map,
            )
            ws[f"AN{row_idx}"] = get_rank(
                rank_row["bridge_holiday_sum_rank"],
                bridge_holiday_priority_rank_map,
            )
            ws[f"AO{row_idx}"] = get_rank(
                rank_row["gw_sum_rank"],
                gw_priority_rank_map,
            )
            ws[f"AP{row_idx}"] = get_rank(
                rank_row["obon_sum_rank"],
                obon_priority_rank_map,
            )
            ws[f"AQ{row_idx}"] = get_rank(
                rank_row["new_year_sum_rank"],
                new_year_priority_rank_map,
            )
            ws[f"AR{row_idx}"] = get_rank(
                rank_row["year_end_sum_rank"],
                year_end_priority_rank_map,
            )
            ws[f"AS{row_idx}"] = get_rank(
                rank_row["black_friday_sum_rank"],
                black_friday_priority_rank_map,
            )

    def build(self) -> bytes:
        zip_buffer = BytesIO()

        regional_facility_map: dict[str, list[FacilityDetail]] = {}

        for facility_detail in self.facility_details:
            key = facility_detail.regional_office

            if key not in regional_facility_map:
                regional_facility_map[key] = []

            regional_facility_map[key].append(facility_detail)

        regional_facility_daily_target_detail_map: dict[
            str, list[FacilityDailyTargetDetail]
        ] = {}

        for facility_daily_target_detail in self.facility_daily_target_details:
            key = facility_daily_target_detail.regional_office

            if key not in regional_facility_daily_target_detail_map:
                regional_facility_daily_target_detail_map[key] = []

            regional_facility_daily_target_detail_map[key].append(
                facility_daily_target_detail
            )

        def add_workbook_to_zip(
            wb: Workbook,
            zip_file: ZipFile,
            *,
            constraint_detail: ConstraintDetail,
            is_regional_version: bool = True,
        ) -> None:
            regional_office = constraint_detail.regional_office

            facility_details = (
                regional_facility_map[regional_office]
                if is_regional_version
                else self.facility_details
            )

            daily_target_details = (
                regional_facility_daily_target_detail_map[regional_office]
                if is_regional_version
                else self.facility_daily_target_details
            )

            output_filename = (
                f"{regional_office}.xlsx" if is_regional_version else "_全国版.xlsx"
            )

            try:
                self._write_common_constraint_sheet(
                    wb[self.COMMON_CONSTRAINT_SHEET_NAME], constraint_detail
                )

                if is_regional_version:
                    regional_office_constraint_ws = wb[
                        self.REGIONAL_OFFICE_CONSTRAINT_SHEET_NAME
                    ]
                    wb.remove(regional_office_constraint_ws)
                else:
                    self._write_regional_office_constraint_sheet(
                        wb[self.REGIONAL_OFFICE_CONSTRAINT_SHEET_NAME]
                    )

                self._write_facility_output_sheet(
                    wb[self.COPILOT_OUTPUT_SHEET_NAME],
                    facility_details,
                )

                self._write_date_header(
                    wb[self.COPILOT_OUTPUT_SHEET_NAME], start_col=18
                )

                for sheet_name in self.FACILITY_OUTPUT_SHEET_NAMES:
                    ws = wb[sheet_name]

                    self._write_facility_output_sheet(ws, facility_details)

                    self._write_date_header(ws)

                self._write_facility_daily_target_sheet(
                    wb[self.FACILITY_DAILY_TARGET_SHEET_NAME],
                    daily_target_details,
                )

                self._write_facility_priority_sheet(
                    wb[self.FACILITY_PRIORITY_SHEET_NAME],
                    facility_details,
                )

                with BytesIO() as workbook_buffer:
                    wb.save(workbook_buffer)
                    zip_file.writestr(
                        output_filename,
                        workbook_buffer.getvalue(),
                    )
            finally:
                wb.close()

        zip_buffer = BytesIO()
        with ZipFile(
            zip_buffer,
            mode="w",
            compression=ZIP_DEFLATED,
        ) as zip_file:
            total_daily_event_limit = 0
            total_target_actual = 0
            total_constraint_cost = 0

            # 支社別版のExcelを作成する
            for constraint_detail in self.constraint_details:
                wb = load_workbook(BytesIO(self.template_workbook_bytes))
                add_workbook_to_zip(wb, zip_file, constraint_detail=constraint_detail)

                total_daily_event_limit += constraint_detail.daily_event_limit
                total_target_actual += constraint_detail.target_actual
                total_constraint_cost += constraint_detail.constraint_cost

            # 全支社版のExcelを作成
            wb = load_workbook(BytesIO(self.template_workbook_bytes))
            add_workbook_to_zip(
                wb,
                zip_file,
                constraint_detail=ConstraintDetail(
                    regional_office="",
                    proposal_period=self.constraint_details[0].proposal_period,
                    daily_event_limit=total_daily_event_limit,
                    weekday_pattern="",
                    target_actual=total_target_actual,
                    constraint_cost=total_constraint_cost,
                ),
                is_regional_version=False,
            )

        return zip_buffer.getvalue()
