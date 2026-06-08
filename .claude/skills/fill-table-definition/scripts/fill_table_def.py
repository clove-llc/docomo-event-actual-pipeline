"""テーブル定義書（Excel）の各テーブルシートを、dbtのsource定義とBigQueryカタログから自動で埋める。

【最重要・絶対遵守】
  BigQueryのテーブルは絶対に上書き・作成・変更・削除しない。許可される操作はデータの確認(参照)のみ。
  本スクリプトはローカルの yml / catalog.json を読み、Excel を書き出すだけで、BigQuery への書き込みは一切行わない。


入力:
  --template : テンプレート xlsx（`tmp_...` という名前の定義シート雛形を含む）
  --sources  : dbtの source 定義 yml（__sources.yml）
  --catalog  : dbt docs generate が出力する catalog.json（BigQuery実型）
  --output   : 出力 xlsx

挙動:
  - source yml の各テーブルごとに、テンプレートの `tmp_...` シートを複製してスタイルを引き継ぐ
  - テーブルレベル（物理テーブル名 / データベース名 / スキーマ名 / 用途・概要）を記入
  - カラムレベル（物理カラム名 / 論理名・説明 / データ型 / NOT NULL / UK / PK / FK・参照先
    / サンプル値 / クレンジング仕様 / 備考）を記入
  - データ型は catalog.json（BigQuery実型）を優先、無ければ空欄
  - 雛形(`tmp_...`)シートは出力から除外（--keep-template で残す）

マッピング規則:
  NOT NULL = not_null テストあり → ○
  UK       = unique テストあり   → ○
  PK       = not_null かつ unique → ○
  FK・参照先 = relationships テストの to / field
  サンプル値 = accepted_values テストの values（あれば）
"""

import argparse
import json
import math
import os
from copy import copy

import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def parse_tests(col):
    """カラム定義のテスト一覧から (not_null, unique, fk文字列, accepted_values) を抽出。"""
    tests = col.get("tests") or col.get("data_tests") or []
    not_null = unique = False
    fk = ""
    accepted = []
    for t in tests:
        if isinstance(t, str):
            if t == "not_null":
                not_null = True
            elif t == "unique":
                unique = True
        elif isinstance(t, dict):
            for key, val in t.items():
                if key == "not_null":
                    not_null = True
                elif key == "unique":
                    unique = True
                elif key == "relationships" and isinstance(val, dict):
                    to = str(val.get("to", "")).strip()
                    field = str(val.get("field", "")).strip()
                    fk = f"{to} / {field}".strip(" /")
                elif key == "accepted_values" and isinstance(val, dict):
                    accepted = list(val.get("values", []) or [])
    return not_null, unique, fk, accepted


def load_sources(path):
    """source yml を {table_name: {description, columns:[...]}} の形に正規化。"""
    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    tables = []
    for src in data.get("sources", []):
        schema = src.get("schema", "")
        database = src.get("database", "")
        for tbl in src.get("tables", []):
            cols = []
            for c in tbl.get("columns", []):
                nn, uk, fk, accepted = parse_tests(c)
                cols.append({
                    "name": c["name"],
                    "description": c.get("description", ""),
                    "not_null": nn,
                    "unique": uk,
                    "pk": nn and uk,
                    "fk": fk,
                    "accepted": accepted,
                })
            tables.append({
                "name": tbl["name"],
                "description": tbl.get("description", ""),
                "schema": schema,
                "database": database,
                "columns": cols,
            })
    return tables


def _columns_from_yml(node):
    cols = []
    for c in node.get("columns", []):
        nn, uk, fk, accepted = parse_tests(c)
        cols.append({
            "name": c["name"],
            "description": c.get("description", ""),
            "not_null": nn,
            "unique": uk,
            "pk": nn and uk,
            "fk": fk,
            "accepted": accepted,
        })
    return cols


def load_models(paths):
    """dbtモデルyml（`models:` ブロック）を読み、tablesと同じ形に正規化する。

    paths: yml ファイル または ディレクトリ（配下の *.yml を再帰探索）のリスト。
    schema / database は yml に無いので空（catalog から補完される）。
    """
    import glob

    files = []
    for p in paths:
        if os.path.isdir(p):
            files += sorted(glob.glob(os.path.join(p, "**", "*.yml"), recursive=True))
            files += sorted(glob.glob(os.path.join(p, "**", "*.yaml"), recursive=True))
        elif os.path.isfile(p):
            files.append(p)
    tables = []
    seen = set()
    for f in files:
        with open(f, encoding="utf-8") as fh:
            data = yaml.safe_load(fh) or {}
        for mdl in data.get("models", []):
            if mdl["name"] in seen:
                continue
            seen.add(mdl["name"])
            tables.append({
                "name": mdl["name"],
                "description": mdl.get("description", ""),
                "schema": "",
                "database": "",
                "columns": _columns_from_yml(mdl),
            })
    return tables


def load_catalog(path):
    """catalog.json から {table_name: {database, schema, types:{col:type}}} を作る。"""
    catalog = {}
    if not path:
        return catalog
    try:
        with open(path, encoding="utf-8") as f:
            cat = json.load(f)
    except FileNotFoundError:
        return catalog
    for group in ("sources", "nodes"):  # sources=raw, nodes=モデル(stg/int/mart)
        for node in cat.get(group, {}).values():
            meta = node.get("metadata", {})
            name = meta.get("name")
            if not name:
                continue
            catalog[name] = {
                "database": meta.get("database", ""),
                "schema": meta.get("schema", ""),
                "types": {cn: cv.get("type", "") for cn, cv in node.get("columns", {}).items()},
            }
    return catalog


# ===== 補助メタ（コード側で管理する設定。yml に項目が無い情報を補う）=====

# 論理テーブル名（出典: main.py の run_pipeline(name=...) 等）
LOGICAL_NAMES = {
    "raw_facility_master": "施設マスタ",
    "raw_facility_daily_deviation_zscore": "施設・日付フラグ別 偏差値Zスコアマスタ",
    "raw_facility_foot_traffic_avg_and_decile_by_flag": "施設・日付フラグ別 人流平均・デシルランクマスタ",
    "raw_date_master": "日付マスタ",
    "raw_date_master_2025_2026": "日付マスタ（2025-2026）",
    "raw_date_master_2026_2027": "日付マスタ（2026-2027）",
    "raw_facility_name_mappings": "施設名 表記ゆれマッピング",
    "raw_facility_actuals": "実績データ（施設別・縦持ち）",
    "raw_venue_performance": "会場別実績データ",
}

# クレンジング仕様（出典: README の日付実績クレンジング規則）
_ACTUAL_CLEANSING = (
    "前後空白をTRIM。＠/@・中止・確認中→NULL、なし→0。完全未入力行は除外。NULL許容Int64。"
)
CLEANSING = {
    ("raw_facility_actuals", "actual_value"): _ACTUAL_CLEANSING,
    ("raw_venue_performance", "daily_result"): _ACTUAL_CLEANSING,
    ("raw_venue_performance", "daily_result_raw"): "正規化前の生値（クレンジング適用前）。",
}

# 複合PK候補（単一PKが無いテーブル。BigQueryで一意性を検証してから採用）
CANDIDATE_PKS = {
    "raw_facility_daily_deviation_zscore": ["facility_code", "date"],
}

# FK関係（子テーブル.カラム → 親テーブル.カラム。BigQueryで包含を検証してから採用）
FK_RULES = {
    ("raw_facility_daily_deviation_zscore", "facility_code"):
        ("raw_facility_master", "facility_code"),
    ("raw_facility_foot_traffic_avg_and_decile_by_flag", "facility_code"):
        ("raw_facility_master", "facility_code"),
}

ENUM_MAX = 12  # distinct がこの数以下のカラムは、サンプル値に全列挙する

# レイヤー別のシートタブ色（スキーマ名から判定）
LAYER_COLORS = {
    "raw": "ED7D31",   # オレンジ
    "int": "00B0F0",   # 水色
    "mart": "7030A0",  # 紫
    "stg": "A6A6A6",   # グレー（参考）
}


def layer_of(schema):
    """スキーマ名からレイヤー種別を判定する（BQ: *_raw / Snowflake: RAW 等の両対応）。"""
    s = (schema or "").lower()
    if s == "raw" or s.endswith("_raw"):
        return "raw"
    if s == "int" or s.endswith("_intermediate"):
        return "int"
    if s == "mart" or s.endswith("_mart"):
        return "mart"
    if s == "stg" or s.endswith("_staging"):
        return "stg"
    return ""


def profile_table(client, project, schema, table, cols):
    """BigQueryへ読み取り(SELECT)のみで、各カラムのNULL/一意性/カーディナリティを実測する。"""
    fq = f"`{project}.{schema}.{table}`"
    sel = ["COUNT(*) AS _total"]
    for col, _ in cols:
        sel.append(f"COUNTIF(`{col}` IS NULL) AS `n__{col}`")
        sel.append(f"COUNT(DISTINCT `{col}`) AS `d__{col}`")
        sel.append(f"CAST(MIN(`{col}`) AS STRING) AS `s__{col}`")  # 非NULLの一例
    row = list(client.query(f"SELECT {', '.join(sel)} FROM {fq}").result())[0]
    total = row["_total"]
    prof = {"_total": total}
    low = []
    for col, typ in cols:
        nulls = row[f"n__{col}"]
        dist = row[f"d__{col}"]
        nonnull = total - nulls
        prof[col] = {
            "type": typ, "nulls": nulls, "dist": dist, "total": total,
            "nn": nulls == 0 and total > 0,
            "uk": dist == nonnull and nonnull > 0,
            "pk": nulls == 0 and dist == total and total > 0,
            "values": None,
            "sample": row[f"s__{col}"],  # 代表値（全カラムに必ず1つ）
        }
        if 0 < dist <= ENUM_MAX:
            low.append(col)
    # 低カーディナリティ列の値を列挙（サンプル値用）
    if low:
        sel2 = [f"ARRAY_AGG(DISTINCT CAST(`{c}` AS STRING) IGNORE NULLS LIMIT {ENUM_MAX + 1}) AS `v__{c}`"
                for c in low]
        r2 = list(client.query(f"SELECT {', '.join(sel2)} FROM {fq}").result())[0]
        for c in low:
            prof[c]["values"] = sorted(r2[f"v__{c}"])
    return prof


def check_composite_pk(client, project, schema, table, combo, total):
    """複合キーが一意かをBigQueryで検証。"""
    concat = ", '|', ".join(f"IFNULL(CAST(`{c}` AS STRING), '∅')" for c in combo)
    q = (f"SELECT COUNT(DISTINCT CONCAT({concat})) AS d "
         f"FROM `{project}.{schema}.{table}`")
    d = list(client.query(q).result())[0]["d"]
    return d == total


def check_fk(client, project, schema, child, ccol, parent, pcol):
    """子のカラム値が親に全て存在するか（孤児ゼロか）をBigQueryで検証。"""
    q = (f"SELECT COUNT(*) AS orphans FROM `{project}.{schema}.{child}` c "
         f"WHERE c.`{ccol}` IS NOT NULL AND NOT EXISTS "
         f"(SELECT 1 FROM `{project}.{schema}.{parent}` p WHERE p.`{pcol}` = c.`{ccol}`)")
    return list(client.query(q).result())[0]["orphans"] == 0


def build_profiles(project, catalog):
    """全テーブルをプロファイルし、複合PK・FKの検証結果も付与して返す。

    返り値: {table: {"cols": prof, "composite_pk": set(...), "fk": {col: "parent.col"}}}
    """
    from google.cloud import bigquery  # 遅延import（--no-profile時は不要）

    client = bigquery.Client(project=project)
    out = {}
    for table, meta in catalog.items():
        schema = meta["schema"]
        cols = [(c, t) for c, t in meta["types"].items()]
        try:
            prof = profile_table(client, project, schema, table, cols)
            total = prof["_total"]
            # 複合PK
            composite = set()
            has_single_pk = any(v.get("pk") for k, v in prof.items() if k != "_total")
            if not has_single_pk and table in CANDIDATE_PKS:
                combo = CANDIDATE_PKS[table]
                if all(c in prof for c in combo) and check_composite_pk(
                        client, project, schema, table, combo, total):
                    composite = set(combo)
            # FK
            fk = {}
            for (ct, ccol), (pt, pcol) in FK_RULES.items():
                if ct == table and check_fk(client, project, schema, table, ccol, pt, pcol):
                    fk[ccol] = f"{pt}.{pcol}"
            out[table] = {"cols": prof, "composite_pk": composite, "fk": fk}
        except Exception as e:  # noqa: BLE001
            print(f"[警告] {table} のプロファイリングをスキップ: {e}")
    return out


# ===== Snowflake モード（BQ ではなく Snowflake を参照して型・プロファイルを取得）=====
# 記述・テスト・論理名は BQ の yml を流用（SFは BQ SQL ベースで作成済み・列の意味は同一）。
# DB/スキーマ・データ型・NULL/一意性・サンプル値は Snowflake から取得する（読み取り SELECT のみ）。
SF_SCHEMA_OF_LAYER = {"raw": "RAW", "stg": "STG"}


def sf_connect(secrets_path):
    """secrets.toml（[snowflake]）からキーペア接続を張る。(conn, database) を返す。"""
    import tomllib

    from cryptography.hazmat.primitives import serialization
    import snowflake.connector

    with open(secrets_path, "rb") as f:
        cfg = tomllib.load(f)["snowflake"]
    with open(cfg["private_key_file"], "rb") as f:
        pk = serialization.load_pem_private_key(f.read(), password=None)
    pkb = pk.private_bytes(serialization.Encoding.DER,
                           serialization.PrivateFormat.PKCS8,
                           serialization.NoEncryption())
    db = cfg.get("database", "HARATO")
    conn = snowflake.connector.connect(
        account=cfg["account"], user=cfg["user"], role=cfg.get("role"),
        warehouse=cfg.get("warehouse"), database=db, private_key=pkb)
    return conn, db


def bq_to_sf_type(bqtype):
    """BigQuery 型を Snowflake 表記に変換（SF未作成の int/mart を SF型風に揃えるため）。"""
    t = (bqtype or "").upper().strip()
    base = t.split("(")[0]
    m = {
        "STRING": "VARCHAR", "BYTES": "BINARY",
        "INT64": "NUMBER(38,0)", "INTEGER": "NUMBER(38,0)",
        "FLOAT64": "FLOAT", "FLOAT": "FLOAT",
        "NUMERIC": "NUMBER(38,9)", "DECIMAL": "NUMBER(38,9)", "BIGNUMERIC": "NUMBER(38,9)",
        "BOOL": "BOOLEAN", "BOOLEAN": "BOOLEAN",
        "DATE": "DATE", "DATETIME": "TIMESTAMP_NTZ", "TIMESTAMP": "TIMESTAMP_NTZ", "TIME": "TIME",
    }
    return m.get(base, t)


def _sf_type(data_type, prec, scale, length):
    """INFORMATION_SCHEMA の型情報を Snowflake の表記に整える。"""
    dt = (data_type or "").upper()
    if dt in ("NUMBER", "DECIMAL", "NUMERIC"):
        if prec is not None:
            return f"NUMBER({prec},{scale or 0})"
        return "NUMBER"
    if dt in ("TEXT", "VARCHAR", "STRING", "CHAR", "CHARACTER"):
        return "VARCHAR"
    return dt or ""


def load_sf_catalog(conn, db, want):
    """want: {table_name(小文字): 'raw'|'stg'} → catalog[name]={database,schema,types{lower:type},_real{lower:actual}}。"""
    cur = conn.cursor()
    cat = {}
    for name, layer in want.items():
        schema = SF_SCHEMA_OF_LAYER[layer]
        cur.execute(
            f"select column_name, data_type, numeric_precision, numeric_scale, character_maximum_length "
            f"from {db}.INFORMATION_SCHEMA.COLUMNS "
            f"where table_schema = %s and table_name = %s order by ordinal_position",
            (schema, name.upper()))
        types, real = {}, {}
        for cn, dt, p, s, ln in cur.fetchall():
            key = cn.lower()
            types[key] = _sf_type(dt, p, s, ln)
            real[key] = cn
        cat[name] = {"database": db, "schema": schema, "types": types, "_real": real}
    cur.close()
    return cat


def _q(ident):
    """Snowflake 識別子を二重引用符でクォート（小文字クォート列・大文字列の両対応）。"""
    return '"' + ident.replace('"', '""') + '"'


def profile_table_sf(conn, db, schema, table, cols_real):
    """Snowflake を読み取り(SELECT)のみで NULL/一意性/サンプル値を実測。cols_real=[(lower, actual), ...]。"""
    cur = conn.cursor()
    fq = f"{db}.{schema}.{_q(table.upper())}"
    sel = ['COUNT(*) AS "_total"']
    for low, actual in cols_real:
        c = _q(actual)
        sel.append(f'COUNT_IF({c} IS NULL) AS "n__{low}"')
        sel.append(f'COUNT(DISTINCT {c}) AS "d__{low}"')
        sel.append(f'MIN(TO_VARCHAR({c})) AS "s__{low}"')  # 非NULLの一例（型に依存しないよう文字列化）
    cur.execute(f"SELECT {', '.join(sel)} FROM {fq}")
    names = [d[0].lower() for d in cur.description]
    row = dict(zip(names, cur.fetchone()))
    total = row["_total"]
    prof = {"_total": total}
    low_cards = []
    for low, actual in cols_real:
        nulls = row[f"n__{low}"] or 0
        dist = row[f"d__{low}"] or 0
        nonnull = total - nulls
        prof[low] = {
            "type": "", "nulls": nulls, "dist": dist, "total": total,
            "nn": nulls == 0 and total > 0,
            "uk": dist == nonnull and nonnull > 0,
            "pk": nulls == 0 and dist == total and total > 0,
            "values": None, "sample": row[f"s__{low}"],
        }
        if 0 < dist <= ENUM_MAX:
            low_cards.append((low, actual))
    for low, actual in low_cards:  # 低カーディナリティ列はサンプル値に全列挙
        c = _q(actual)
        cur.execute(f"SELECT DISTINCT TO_VARCHAR({c}) v FROM {fq} "
                    f"WHERE {c} IS NOT NULL ORDER BY 1 LIMIT {ENUM_MAX + 1}")
        prof[low]["values"] = [r[0] for r in cur.fetchall()]
    cur.close()
    return prof


def check_composite_pk_sf(conn, db, schema, table, combo_real, total):
    """複合キーが一意かを Snowflake で検証。combo_real=実カラム名のリスト。"""
    cur = conn.cursor()
    concat = ", '|', ".join(f"NVL(TO_VARCHAR({_q(c)}), '∅')" for c in combo_real)
    cur.execute(f"SELECT COUNT(DISTINCT CONCAT({concat})) FROM {db}.{schema}.{_q(table.upper())}")
    d = cur.fetchone()[0]
    cur.close()
    return d == total


def build_profiles_sf(conn, db, sf_catalog):
    """全テーブルを Snowflake でプロファイルし、複合PK・FK（既知ルール）も付与して返す。"""
    out = {}
    for table, meta in sf_catalog.items():
        schema = meta["schema"]
        real = meta["_real"]
        cols_real = [(low, real[low]) for low in meta["types"]]
        try:
            prof = profile_table_sf(conn, db, schema, table, cols_real)
            total = prof["_total"]
            composite = set()
            has_single_pk = any(v.get("pk") for k, v in prof.items() if k != "_total")
            if not has_single_pk and table in CANDIDATE_PKS:
                combo = CANDIDATE_PKS[table]
                if all(c in real for c in combo) and check_composite_pk_sf(
                        conn, db, schema, table, [real[c] for c in combo], total):
                    composite = set(combo)
            # FK は既知ルール（BQで包含検証済み・データは SF と一致）を documented として付与
            fk = {ccol: f"{pt}.{pcol}" for (ct, ccol), (pt, pcol) in FK_RULES.items() if ct == table}
            out[table] = {"cols": prof, "composite_pk": composite, "fk": fk}
        except Exception as e:  # noqa: BLE001
            print(f"[警告] {table} のSnowflakeプロファイルをスキップ: {e}")
    return out


def find_template_sheet(wb):
    """名前が tmp で始まる定義シート雛形を返す。"""
    for ws in wb.worksheets:
        if ws.title.lower().startswith("tmp"):
            return ws
    raise SystemExit("テンプレート定義シート（tmp...）が見つかりません。")


def introspect(ws):
    """雛形シートから、見出し行・列マップ・データ行スタイル・凡例位置を読み取る。"""
    header_row = None
    header_map = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "No" and cell.column == 1:
                header_row = cell.row
                break
        if header_row:
            break
    if header_row is None:
        raise SystemExit("カラム見出し行（No）が見つかりません。")

    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            header_map[str(v).strip()] = c

    # データ行スタイル（奇数/偶数の2種）を退避
    style_rows = {}
    for variant, r in (("odd", header_row + 1), ("even", header_row + 2)):
        style_rows[variant] = {
            c: {
                "font": copy(ws.cell(row=r, column=c).font),
                "fill": copy(ws.cell(row=r, column=c).fill),
                "border": copy(ws.cell(row=r, column=c).border),
                "alignment": copy(ws.cell(row=r, column=c).alignment),
            }
            for c in range(1, max_col + 1)
        }
    data_row_height = ws.row_dimensions[header_row + 1].height

    # 凡例（【記入要領】）の開始行
    legend_start = None
    for row in ws.iter_rows(min_row=header_row + 1):
        cell = row[0]
        if isinstance(cell.value, str) and cell.value.startswith("【"):
            legend_start = cell.row
            break

    # 雛形が持つデータ行数を実測（罫線付きの連続行を数える）
    existing = 0
    r = header_row + 1
    while r <= ws.max_row:
        if legend_start and r >= legend_start:
            break
        side = ws.cell(row=r, column=1).border.left
        if side and side.style:
            existing += 1
            r += 1
        else:
            break

    return {
        "header_row": header_row,
        "header_map": header_map,
        "max_col": max_col,
        "style_rows": style_rows,
        "data_row_height": data_row_height,
        "legend_start": legend_start,
        "existing": existing,
    }


def fill_meta(ws, desired):
    """テーブルレベルのラベルセルを探して右隣に値を書く。

    1セルに複数ラベルが改行で結合されている場合（例: 'データベース名\\nスキーマ名'）は、
    各値も同じ並び順で改行結合して書き込む。
    """
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            parts = [p.strip() for p in cell.value.split("\n") if p.strip()]
            if not parts or not all(p in desired for p in parts):
                continue  # ラベルセルでない（値セルやタイトル等）
            value = "\n".join(str(desired[p]) for p in parts)
            ws.cell(row=cell.row, column=cell.column + 1, value=value)


def set_title(ws, table_name):
    """{テーブル名} プレースホルダを実テーブル名に置換。"""
    for row in ws.iter_rows(max_row=2):
        for cell in row:
            if isinstance(cell.value, str) and "テーブル名" in cell.value and "{" in cell.value:
                cell.value = table_name
                return


def apply_style(cell, style):
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])


def fill_sheet(ws, table, cat, info, profile=None):
    col_types = cat.get("types", {})
    prof_cols = (profile or {}).get("cols", {})
    composite_pk = (profile or {}).get("composite_pk", set())
    fk_map = (profile or {}).get("fk", {})
    tname = table["name"]

    set_title(ws, tname)
    fill_meta(ws, {
        "物理テーブル名": tname,
        "論理テーブル名": LOGICAL_NAMES.get(tname, ""),
        "データベース名": cat.get("database") or table["database"],
        "スキーマ名": cat.get("schema") or table["schema"],
        "更新タイミング": "",
        "用途・概要": table["description"],
    })

    hr = info["header_row"]
    hmap = info["header_map"]
    max_col = info["max_col"]
    n = len(table["columns"])

    # データ行数を列数に合わせて調整（凡例を押し下げ / 余剰行を削除）
    existing = info["existing"]
    if n > existing:
        ws.insert_rows(hr + 1 + existing, amount=n - existing)
    elif n < existing:
        ws.delete_rows(hr + 1 + n, amount=existing - n)

    for i, col in enumerate(table["columns"], start=1):
        r = hr + i
        variant = info["style_rows"]["odd" if i % 2 == 1 else "even"]
        for c in range(1, max_col + 1):
            apply_style(ws.cell(row=r, column=c), variant[c])
        if info["data_row_height"]:
            ws.row_dimensions[r].height = info["data_row_height"]

        cname = col["name"]
        p = prof_cols.get(cname, {})
        # yml テスト と BigQuery 実測 を統合（OR）
        not_null = col["not_null"] or p.get("nn", False)
        unique = col["unique"] or p.get("uk", False)
        is_pk = col["pk"] or p.get("pk", False) or (cname in composite_pk)
        fk = col["fk"] or fk_map.get(cname, "")
        # サンプル値: yml accepted_values 優先 → 低カーディナリティは全列挙 → それ以外は代表値1つ
        if col["accepted"]:
            sample = " / ".join(map(str, col["accepted"]))
        elif p.get("values"):
            vals = p["values"]
            sample = " / ".join(map(str, vals[:ENUM_MAX]))
            if len(vals) > ENUM_MAX:
                sample += " …"
        elif p.get("sample") is not None:
            sample = str(p["sample"])  # 高カーディナリティ列も必ず1例を入れる
        else:
            sample = ""

        # 備考: NULL所見を自動記入（全行NULL→未投入 / 部分NULL→件数・割合）
        note = ""
        if p:
            nulls, total = p.get("nulls", 0), p.get("total", 0)
            if total and nulls == total:
                note = "全行NULL（未投入）"
            elif nulls:
                pct = nulls / total * 100
                pct_s = "<1%" if round(pct) == 0 else f"{round(pct)}%"
                note = f"NULL {nulls:,}件 ({pct_s})"

        values = {
            "No": i,
            "物理カラム名": cname,
            "論理名・説明": col["description"],
            "データ型": col_types.get(cname, "") or p.get("type", ""),
            "NOT NULL": "○" if not_null else "",
            "UK": "○" if unique else "",
            "PK": "○" if is_pk else "",
            "FK・参照先": fk,
            "サンプル値": sample,
            "クレンジング仕様": CLEANSING.get((tname, cname), ""),
            "備考": note,
        }
        for header, c in hmap.items():
            if header in values:
                ws.cell(row=r, column=c, value=values[header])

    # 値の無い末尾行（雛形由来の空の罫線行など）を除去して整える
    last = 0
    for rr in range(1, ws.max_row + 1):
        if any(ws.cell(row=rr, column=c).value not in (None, "") for c in range(1, max_col + 1)):
            last = rr
    if ws.max_row > last:
        ws.delete_rows(last + 1, ws.max_row - last)


def _col_width(ws, idx):
    w = ws.column_dimensions[get_column_letter(idx)].width
    return w if w else 8.43


def _disp_width(s):
    """全角=2 / 半角=1 で表示幅を概算。"""
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in s)


def _lines_needed(text, avail):
    """折り返し時に必要な行数（改行・列幅から概算）。"""
    total = 0
    for ln in str(text).split("\n"):
        total += max(1, math.ceil(_disp_width(ln) / max(avail, 1)))
    return max(total, 1)


def autofit_rows(ws):
    """文字が隠れないよう、内容・列幅・結合・フォントサイズから行高を自動調整する。

    併せて、内容のあるセルは折り返し(wrap_text)を有効化する。既存の行高より低くはしない。
    """
    merges = list(ws.merged_cells.ranges)

    def merge_at(r, c):
        for m in merges:
            if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
                return m
        return None

    for row in ws.iter_rows():
        row_idx = row[0].row
        best = 0.0
        for cell in row:
            if cell.value in (None, ""):
                continue
            r, c = cell.row, cell.column
            # A1（左上の見出し）は折り返さず1行表示（枠はみ出し可）。高さ計算からも除外。
            if r == 1 and c == 1:
                al = cell.alignment
                if al.wrap_text:
                    cell.alignment = Alignment(
                        horizontal=al.horizontal, vertical=al.vertical,
                        wrap_text=False, text_rotation=al.text_rotation, indent=al.indent,
                    )
                continue
            m = merge_at(r, c)
            if m and (r != m.min_row or c != m.min_col):
                continue  # 結合セルの先頭以外はスキップ
            if m:
                avail = sum(_col_width(ws, cc) for cc in range(m.min_col, m.max_col + 1))
                rowspan = m.max_row - m.min_row + 1
            else:
                avail = _col_width(ws, c)
                rowspan = 1
            avail = max(avail - 1.0, 1.0)  # 余白分を引いて安全側に
            lines = _lines_needed(cell.value, avail)
            font_size = cell.font.size or 11
            line_h = font_size * 1.35
            need = (lines * line_h) / rowspan + 4.0  # 上下パディング
            best = max(best, min(need, 409.0))  # Excelの行高上限(409pt)を超えない

            al = cell.alignment
            if not al.wrap_text:
                cell.alignment = Alignment(
                    horizontal=al.horizontal, vertical=al.vertical or "top",
                    wrap_text=True, text_rotation=al.text_rotation, indent=al.indent,
                )
        if best:
            current = ws.row_dimensions[row_idx].height or 0
            ws.row_dimensions[row_idx].height = max(best, current, 15.0)


def rebuild_table_list(ws, entries, link_font=None):
    """「テーブル一覧」シートを全テーブルで再構築する。

    entries: [(物理テーブル名, 論理名, 用途・概要, レイヤー, シート名), ...]
    既存のヘッダー行とデータ行スタイルを流用し、行を過不足なく書き直す。
    「リンク」列があれば、各定義シートへ遷移する内部ハイパーリンク（表示文字「テーブル定義書」）を入れる。
    """
    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "物理テーブル名":
                header_row = cell.row
                break
        if header_row:
            break
    if header_row is None:
        return
    hmap = {ws.cell(row=header_row, column=c).value: c
            for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value}
    no_col = next((c for c in range(1, ws.max_column + 1)
                   if ws.cell(row=header_row, column=c).value == "No"), 1)
    max_col = ws.max_column

    # データ行スタイル(奇/偶)を退避
    styles = {}
    for variant, rr in (("odd", header_row + 1), ("even", header_row + 2)):
        styles[variant] = {c: {
            "font": copy(ws.cell(row=rr, column=c).font),
            "fill": copy(ws.cell(row=rr, column=c).fill),
            "border": copy(ws.cell(row=rr, column=c).border),
            "alignment": copy(ws.cell(row=rr, column=c).alignment),
        } for c in range(1, max_col + 1)}

    # 既存データ行（罫線付き連続行）を数える
    existing = 0
    r = header_row + 1
    while r <= ws.max_row:
        side = ws.cell(row=r, column=1).border.left
        if side and side.style:
            existing += 1
            r += 1
        else:
            break
    n = len(entries)
    if n > existing:
        ws.insert_rows(header_row + 1 + existing, amount=n - existing)
    elif n < existing:
        ws.delete_rows(header_row + 1 + n, amount=existing - n)

    link_col = hmap.get("リンク")
    for i, entry in enumerate(entries, start=1):
        phys, logical, summary, _layer, sheet = (list(entry) + [None] * 5)[:5]
        r = header_row + i
        variant = styles["odd" if i % 2 == 1 else "even"]
        for c in range(1, max_col + 1):
            apply_style(ws.cell(row=r, column=c), variant[c])
        vals = {"No": i, "物理テーブル名": phys, "論理名": logical, "用途・概要": summary}
        for header, c in hmap.items():
            if header in vals:
                ws.cell(row=r, column=c, value=vals[header])
        if no_col not in hmap.values():
            ws.cell(row=r, column=no_col, value=i)
        # リンク列: 各定義シートへ遷移する内部ハイパーリンク
        if link_col and sheet:
            cell = ws.cell(row=r, column=link_col, value="テーブル定義書")
            # 他列と同じ罫線・ストライプを適用（テンプレ5列目は無地のため隣列の書式を流用）
            ref = next((variant[c] for c in (link_col - 1, 2, 1) if variant.get(c)), None)
            if ref:
                cell.fill = copy(ref["fill"])
                cell.border = copy(ref["border"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            safe = sheet.replace("'", "''")
            cell.hyperlink = f"#'{safe}'!A1"
            if link_font is not None:
                cell.font = copy(link_font)


def sheet_name_for(table_name, used):
    """Excelのシート名制限(31文字・重複不可)に合わせる。"""
    base = table_name[:31]
    name = base
    k = 1
    while name in used:
        suffix = f"_{k}"
        name = base[: 31 - len(suffix)] + suffix
        k += 1
    used.add(name)
    return name


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template",
                    default=".claude/skills/fill-table-definition/tmp_テーブル定義書.xlsx")
    ap.add_argument("--sources", default="docomo_event/models/staging/__sources.yml")
    ap.add_argument("--catalog", default="docomo_event/target/catalog.json")
    ap.add_argument("--output", default="docs/テーブル定義書.xlsx")
    ap.add_argument("--keep-template", action="store_true",
                    help="雛形(tmp...)シートを出力に残す")
    ap.add_argument("--table", default=None,
                    help="指定したテーブル名のみを出力（お試し実行用）")
    ap.add_argument("--models", nargs="*", default=None,
                    help="raw以外に定義書へ含めるモデルyml/ディレクトリ。"
                         "未指定なら engine 既定（bq=intermediate+marts / sf=staging）。空指定で無効")
    ap.add_argument("--engine", choices=["bq", "sf", "mixed"], default="bq",
                    help="型・プロファイルの参照先。bq=BigQuery / sf=Snowflake / "
                         "mixed=raw,stgはSF・int,martはBQ（BQ型→SF型へ変換）")
    ap.add_argument("--sf-secrets", default="streamlit/uploader/.streamlit/secrets.toml",
                    help="Snowflake 接続情報（[snowflake]）。engine=sf/mixed のとき使用")
    ap.add_argument("--layers", nargs="*", default=None,
                    help="出力レイヤーを限定（例: raw stg）。未指定なら全て")
    ap.add_argument("--db-name", default=None,
                    help="データベース名の表示を上書き（例: docomo_db）。実測スキーマには影響しない")
    ap.add_argument("--schema-as-layer", action="store_true",
                    help="スキーマ名の表示をレイヤー名（raw/stg/int/mart）にする")
    ap.add_argument("--profile", dest="profile", action="store_true", default=True,
                    help="DBを読み取り(SELECT)してNOT NULL/UK/PK/FK/サンプル値を実測（既定: 有効）")
    ap.add_argument("--no-profile", dest="profile", action="store_false",
                    help="DBへのプロファイリングを行わず yml のみで埋める")
    ap.add_argument("--project", default="digital-well-456700-i9",
                    help="プロファイリング対象のGCPプロジェクトID（engine=bq/mixed）")
    args = ap.parse_args()

    # --models の既定を engine 別に解決
    if args.models is None:
        if args.engine == "sf":
            args.models = ["docomo_event/models/staging/_staging__models.yml"]
        elif args.engine == "mixed":
            args.models = ["docomo_event/models/staging/_staging__models.yml",
                           "docomo_event/models/intermediate",
                           "docomo_event/models/marts/planning"]
        else:
            args.models = ["docomo_event/models/intermediate",
                           "docomo_event/models/marts/planning"]

    # BQ catalog は bq/mixed で使用（型・スキーマ（=レイヤー）判定）
    bq_catalog = load_catalog(args.catalog) if args.engine in ("bq", "mixed") else {}

    tables = load_sources(args.sources)
    layers = {t["name"]: "raw" for t in tables}
    if args.models:
        models = load_models(args.models)
        for m in models:
            if args.engine == "sf":
                layers[m["name"]] = "stg"
            elif args.engine == "mixed":
                layers[m["name"]] = layer_of(bq_catalog.get(m["name"], {}).get("schema", "")) or "stg"
            else:
                layers[m["name"]] = "model"
        tables += models

    if args.table:
        tables = [t for t in tables if t["name"] == args.table]
        if not tables:
            raise SystemExit(f"テーブルが見つかりません: {args.table}")
    if args.layers:
        tables = [t for t in tables if layers.get(t["name"]) in args.layers]
        if not tables:
            raise SystemExit(f"指定レイヤーに該当テーブルなし: {args.layers}")

    # 参照先の振り分け: SF=raw/stg、BQ=int/mart（bq/sf エンジンでは一方に寄せる）
    if args.engine == "sf":
        sf_tables, bq_tables = tables, []
    elif args.engine == "bq":
        sf_tables, bq_tables = [], tables
    else:  # mixed
        sf_tables = [t for t in tables if layers[t["name"]] in ("raw", "stg")]
        bq_tables = [t for t in tables if layers[t["name"]] in ("int", "mart")]

    sf_conn = None
    catalog = {}
    profiles = {}

    # SF（raw / stg）: 型・プロファイルを Snowflake から取得（読み取りのみ）
    if sf_tables:
        want = {t["name"]: layers[t["name"]] for t in sf_tables}
        sf_conn, sf_db = sf_connect(args.sf_secrets)
        sfcat = load_sf_catalog(sf_conn, sf_db, want)
        catalog.update(sfcat)
        if args.profile:
            try:
                profiles.update(build_profiles_sf(sf_conn, sf_db, sfcat))
                print(f"Snowflakeプロファイリング: {len(sfcat)}テーブル")
            except Exception as e:  # noqa: BLE001
                print(f"[警告] Snowflakeプロファイリングをスキップ: {e}")

    # BQ（int / mart、または engine=bq の全テーブル）: catalog.json ＋ BQ実測
    if bq_tables:
        if not bq_catalog:
            bq_catalog = load_catalog(args.catalog)
        bqcat = {t["name"]: bq_catalog[t["name"]] for t in bq_tables if t["name"] in bq_catalog}
        if args.engine == "mixed":  # int/mart の BQ型を SF表記へ変換して統一
            for meta in bqcat.values():
                meta["types"] = {c: bq_to_sf_type(t) for c, t in meta["types"].items()}
        catalog.update(bqcat)
        if args.profile:
            try:
                profiles.update(build_profiles(args.project, bqcat))
                print(f"BigQueryプロファイリング: {len(bqcat)}テーブル")
            except Exception as e:  # noqa: BLE001
                print(f"[警告] BigQueryプロファイリングをスキップ: {e}")

    # 表示上書き（プロファイル後なので実測には影響しない）: データベース名 / スキーマ名
    for t in tables:
        name = t["name"]
        if name in catalog:
            if args.db_name:
                catalog[name]["database"] = args.db_name
            if args.schema_as_layer:
                catalog[name]["schema"] = layers[name]

    wb = load_workbook(args.template)
    template_ws = find_template_sheet(wb)
    info = introspect(template_ws)

    used_names = {ws.title for ws in wb.worksheets}
    created = []
    sheet_of = {}   # table_name -> シート名
    for table in tables:
        ws = wb.copy_worksheet(template_ws)
        # copy_worksheet が引き継がない設定を補完
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = template_ws.freeze_panes
        new_name = sheet_name_for(table["name"], used_names)
        ws.title = new_name
        sheet_of[table["name"]] = new_name
        # レイヤー別のタブ色（スキーマ名から判定）
        schema = catalog.get(table["name"], {}).get("schema", "")
        color = LAYER_COLORS.get(layer_of(schema))
        if color:
            ws.sheet_properties.tabColor = color
        fill_sheet(ws, table, catalog.get(table["name"], {}), info,
                   profile=profiles.get(table["name"]))
        created.append(ws)

    if not args.keep_template:
        wb.remove(template_ws)

    # シート順: 表紙 → テーブル一覧 → 各テーブル（→ 雛形）
    front = [ws for ws in wb.worksheets if ws.title in ("表紙", "テーブル一覧")]
    rest = [ws for ws in wb.worksheets if ws not in front and ws not in created
            and ws is not template_ws]
    order = front + created + rest
    if args.keep_template and template_ws in wb.worksheets:
        order += [template_ws]
    wb._sheets = order  # noqa: SLF001

    # テーブル一覧を全テーブル（raw + モデル）で再構築（リンク列付き）
    if "テーブル一覧" in wb.sheetnames and not args.table:
        from openpyxl.styles import Font as _Font
        link_font = _Font(name="Yu Gothic", size=11, color="0563C1", underline="single")
        entries = [(t["name"], LOGICAL_NAMES.get(t["name"], ""), t["description"],
                    layers.get(t["name"], ""), sheet_of.get(t["name"]))
                   for t in tables]
        rebuild_table_list(wb["テーブル一覧"], entries, link_font=link_font)

    # 文字が隠れないよう全シートの行高を自動調整
    for ws in wb.worksheets:
        autofit_rows(ws)

    # 出力先ディレクトリを作成して保存
    out_dir = os.path.dirname(args.output)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(args.output)
    if sf_conn is not None:
        sf_conn.close()
    print(f"生成完了: {args.output}（{len(created)}テーブル分のシートを作成）")


if __name__ == "__main__":
    main()
