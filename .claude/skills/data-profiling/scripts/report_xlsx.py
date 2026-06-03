"""プロファイリング結果をテーブル別シートのExcelブックに描画する（自己完結）。

構成: 表紙 / テーブル一覧 / テーブル別シート（概要・Phase2・Phase3・Phase4・テーブル間整合）。
スタイルは「テーブル定義書」と同体裁（レイヤー別タブ色・リンク列・A1一行・行高オートフィット）。
"""

import math

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- パレット ----
C_HEADER = "305496"
C_META = "D9E1F2"
C_STRIPE = "F2F6FC"
C_LIST = "4472C4"
LAYER_COLORS = {"raw": "ED7D31", "int": "00B0F0", "mart": "7030A0", "stg": "A6A6A6"}

F_TITLE = Font(name="Yu Gothic", size=16, bold=True, color="1F3864")
F_SEC = Font(name="Yu Gothic", size=12, bold=True, color="1F3864")
F_H = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
F_LBL = Font(name="Yu Gothic", size=11, bold=True, color="1F3864")
F_B = Font(name="Yu Gothic", size=11)
F_LINK = Font(name="Yu Gothic", size=11, color="0563C1", underline="single")
F_NOTE = Font(name="Yu Gothic", size=9, color="808080")

FILL_H = PatternFill("solid", fgColor=C_HEADER)
FILL_META = PatternFill("solid", fgColor=C_META)
FILL_STRIPE = PatternFill("solid", fgColor=C_STRIPE)
FILL_LIST = PatternFill("solid", fgColor=C_LIST)

_TH = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)
LT = Alignment(horizontal="left", vertical="top", wrap_text=True)
LC = Alignment(horizontal="left", vertical="center", wrap_text=True)
CC = Alignment(horizontal="center", vertical="center", wrap_text=True)


def layer_of(schema):
    s = (schema or "").lower()
    for suf, key in (("_raw", "raw"), ("_intermediate", "int"), ("_mart", "mart"), ("_staging", "stg")):
        if s.endswith(suf):
            return key
    return ""


# レイヤー順（raw → stg → int → mart）でテーブルを並べる
_LAYER_RANK = {"raw": 0, "stg": 1, "int": 2, "mart": 3}


def sort_key(r):
    return (_LAYER_RANK.get(layer_of(r["dataset"]), 9), r["table"])


def _c(ws, r, col, v=None, font=F_B, fill=None, al=LT, border=True):
    x = ws.cell(row=r, column=col, value=v)
    x.font = font
    if fill:
        x.fill = fill
    x.alignment = al
    if border:
        x.border = BORDER
    return x


def num(x):
    if x is None:
        return ""
    if isinstance(x, bool):
        return str(x)
    if isinstance(x, float):
        return f"{x:,.2f}".rstrip("0").rstrip(".")
    return f"{x:,}" if isinstance(x, int) else str(x)


def sheet_name_for(name, used):
    base = name[:31]
    nm = base
    k = 1
    while nm in used:
        suf = f"_{k}"
        nm = base[: 31 - len(suf)] + suf
        k += 1
    used.add(nm)
    return nm


def _disp_w(s):
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in s)


def autofit(ws):
    """A1は1行（折り返しなし）、他は内容・幅から行高を自動調整。"""
    merges = list(ws.merged_cells.ranges)

    def merge_at(r, c):
        for m in merges:
            if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
                return m
        return None

    for row in ws.iter_rows():
        ridx = row[0].row
        best = 0.0
        for cell in row:
            if cell.value in (None, ""):
                continue
            r, c = cell.row, cell.column
            # A1見出し / 罫線の無い自由テキスト（セクション見出し・箇条書き）は折り返さず1行
            free_text = not (cell.border.left and cell.border.left.style)
            if (r == 1 and c == 1) or free_text:
                a = cell.alignment
                if a.wrap_text:
                    cell.alignment = Alignment(horizontal=a.horizontal, vertical=a.vertical,
                                               wrap_text=False)
                continue
            m = merge_at(r, c)
            if m and (r != m.min_row or c != m.min_col):
                continue
            if m:
                avail = sum((ws.column_dimensions[get_column_letter(cc)].width or 8.43)
                            for cc in range(m.min_col, m.max_col + 1))
            else:
                avail = ws.column_dimensions[get_column_letter(c)].width or 8.43
            avail = max(avail - 1.0, 1.0)
            lines = sum(max(1, math.ceil(_disp_w(ln) / avail)) for ln in str(cell.value).split("\n"))
            fs = cell.font.size or 11
            best = max(best, min(lines * fs * 1.35 + 4.0, 409.0))
        if best:
            cur = ws.row_dimensions[ridx].height or 0
            ws.row_dimensions[ridx].height = max(best, cur, 15.0)


def _quality_lines(quality, cols_out):
    """各品質所見を、平易な説明付きの1行テキストにする。"""
    out = []
    if quality.get("dup_rows"):
        out.append(f"・完全重複行: {num(quality['dup_rows'])}件 … 全カラムの値が完全に一致する行。"
                   "同じデータの二重登録の可能性。")
    if quality.get("sentinels"):
        parts = []
        for c, items in quality["sentinels"].items():
            vals = " / ".join(f"「{v if v else '空白'}」{num(n)}件" for v, n in items)
            parts.append(f"{c}（{vals}）")
        out.append("・エラー値/センチネル: " + "、".join(parts)
                   + " … 本来の値の代わりに、エラー記号や未入力を表す記号が入っている件数"
                   "（例: #N/A はスプレッドシートのエラー、なし/確認中 は手入力の未確定値）。")
    if quality.get("outliers"):
        ex = quality.get("outlier_examples", {})
        parts = []
        for c, n in quality["outliers"].items():
            vs = ex.get(c) or []
            extra = f"（例: {', '.join(num(v) for v in vs)}）" if vs else ""
            parts.append(f"{c}={num(n)}件{extra}")
        out.append("・外れ値(IQR法): " + ", ".join(parts)
                   + " … 四分位範囲(IQR)の1.5倍を超えて離れた、ごく少数の値。入力ミスや特殊ケースの可能性。")
    if quality.get("negatives"):
        parts = []
        for c, n in quality["negatives"].items():
            mn = (cols_out.get(c, {}).get("num") or {}).get("min")
            parts.append(f"{c}={num(n)}件" + (f"（最小 {num(mn)}）" if mn is not None else ""))
        out.append("・負値: " + ", ".join(parts)
                   + " … マイナスの値。数量・金額・件数系では通常あり得ず、入力誤りの可能性。")
    if quality.get("date_gaps"):
        parts = []
        for c, v in quality["date_gaps"].items():
            exs = v.get("examples") or []
            extra = f"（例: {', '.join(exs)}…）" if exs else ""
            parts.append(f"{c}（{num(v['missing'])}日欠損／対象期間{v['span']}日中{v['days']}日のみ存在）{extra}")
        out.append("・日付の欠損日: " + "、".join(parts)
                   + " … 期間内なのにレコードが1件も無い日付。データ未取得・取込漏れの可能性。")
    if quality.get("date_reversed"):
        out.append(f"・開始日>終了日の逆転: {num(quality['date_reversed'])}件 … 終了日が開始日より前のレコード。"
                   "日付の入力誤りの可能性。")
    full_null = [c for c, i in cols_out.items() if i["null_pct"] == 100]
    if full_null:
        out.append("・全行NULL(未投入)列: " + ", ".join(full_null)
                   + " … 全行が空＝一度も値が入っていない列。未実装・未投入の可能性。")
    return out or ["・目立った問題は検出されませんでした。"]


# ---- 表紙 ----
def build_cover(wb, scope, project, gen_date, n_tables):
    ws = wb.active
    ws.title = "表紙"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col, w in zip("BCDE", (22, 16, 16, 44)):
        ws.column_dimensions[col].width = w
    ws.merge_cells("B2:E2")
    t = ws["B2"]
    t.value = "データプロファイリングレポート"
    t.font = F_TITLE
    t.alignment = CC
    ws.row_dimensions[2].height = 40
    meta = [("対象データセット", scope), ("GCPプロジェクト", project),
            ("生成日", gen_date), ("対象テーブル数", str(n_tables)),
            ("調査方式", "BigQuery 読み取り(SELECT)のみ・並列（書き込みなし）")]
    r = 5
    for k, v in meta:
        _c(ws, r, 2, k, font=F_H, fill=FILL_H, al=CC)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        _c(ws, r, 3, v, al=LC)
        for col in (4, 5):
            _c(ws, r, col)
        ws.row_dimensions[r].height = 22
        r += 1


# ---- テーブル一覧 ----
def build_list(wb, results, sheet_of):
    ws = wb.create_sheet("テーブル一覧")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_LIST
    _c(ws, 1, 1, "テーブル一覧", font=F_SEC, border=False)
    ws.row_dimensions[1].height = 26
    heads = [("No", 5), ("物理テーブル名", 46), ("論理名", 26), ("行数", 12),
             ("最終更新", 16), ("主な所見", 50), ("リンク", 14)]
    for i, (h, w) in enumerate(heads, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        _c(ws, 3, i, h, font=F_H, fill=FILL_H, al=CC)
    for i, r in enumerate(sorted(results, key=sort_key), 1):
        rr = 3 + i
        st = FILL_STRIPE if i % 2 == 0 else None
        rows = f"{r['rows']:,}" if isinstance(r["rows"], int) else "?"
        f = " / ".join(r["findings"]) if r["findings"] else "—"
        _c(ws, rr, 1, i, fill=st, al=CC)
        _c(ws, rr, 2, r["table"], fill=st, al=LC)
        _c(ws, rr, 3, r.get("logical", ""), fill=st, al=LC)
        _c(ws, rr, 4, rows, fill=st, al=CC)
        _c(ws, rr, 5, r.get("modified", ""), fill=st, al=CC)
        _c(ws, rr, 6, f, fill=st, al=LC)
        lc = _c(ws, rr, 7, "レポート", fill=st, al=CC)
        lc.font = F_LINK
        sn = sheet_of[(r["dataset"], r["table"])]
        lc.hyperlink = f"#'{sn}'!A1"
        ws.row_dimensions[rr].height = 28
    ws.freeze_panes = ws.cell(row=4, column=1)
    autofit(ws)


# ---- テーブル別シート ----
def build_detail(wb, res, sheet_name, cross_lines):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    color = LAYER_COLORS.get(layer_of(res["meta"].get("schema") or
                                       res["dataset"].split("docomo_event_")[-1]))
    # dataset名からレイヤー推定（schemaが無くてもdataset名で）
    if not color:
        color = LAYER_COLORS.get(layer_of(res["dataset"]))
    if color:
        ws.sheet_properties.tabColor = color

    _c(ws, 1, 1, res["table"], font=F_TITLE, border=False)
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=False)
    ws.row_dimensions[1].height = 26
    # 列: No / 物理カラム名 / 型 / NULL率 / 一意率 / distinct / min / 中央 / max / σ / 区分値 / 品質所見
    widths = [15, 30, 10, 8, 8, 10, 12, 12, 12, 10, 30, 26]
    NCOL = len(widths)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    meta = res["meta"]
    f = meta.get("freq") or {}
    rows_s = f"{res['rows']:,}" if isinstance(res["rows"], int) else "?"
    size = meta.get("size_bytes")
    overview = [
        ("物理テーブル名 / 論理名", f"{res['table']} / {res.get('logical','')}"),
        ("DB.スキーマ / 種別", f"{meta.get('database','')}.{meta.get('schema','')} / {meta.get('typ','?')}"),
        ("行数 / サイズ", rows_s + (f" / {size/1e6:.1f}MB" if size else "")),
        ("作成 / 最終更新", f"{meta.get('created','?')} / {meta.get('modified','?')}"),
        ("更新頻度(180日)", f"書込{f['writes']}回 / 稼働{f['active_days']}日 ({f['first_day']}〜{f['last_day']})"
         if f.get("writes") else "—"),
    ]
    rd = res.get("recent_daily")
    if rd and rd["days"]:
        days_str = " / ".join(f"{d}: {num(n)}件" for d, n in rd["days"])
        overview.append((f"直近{len(rd['days'])}日のレコード数（{rd['col']}）", days_str))
    r = 3
    _c(ws, r, 1, "■ テーブル概要", font=F_SEC, border=False)
    r += 1
    for k, v in overview:
        # ラベルは A:B 結合（広めに取り1行で収める）、値は C 以降に結合
        _c(ws, r, 1, k, font=F_LBL, fill=FILL_META, al=LC)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        _c(ws, r, 2, fill=FILL_META)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=NCOL)
        _c(ws, r, 3, v, al=LC)
        for c in range(4, NCOL + 1):
            _c(ws, r, c)
        ws.row_dimensions[r].height = 22
        r += 1
    r += 1

    # Phase2
    _c(ws, r, 1, "■ Phase2 カラム別プロファイル", font=F_SEC, border=False)
    r += 1
    _c(ws, r, 1, "NULL率=値が空(欠損)の割合 / 一意率=非NULLのうち異なる値の割合(100%＝重複なし) / "
       "distinct=異なる値の数 / 区分値=出現の多い値（括弧内は件数）", font=F_NOTE, border=False)
    r += 1
    ch = ["No", "物理カラム名", "型", "NULL率", "一意率", "distinct",
          "min", "中央", "max", "σ", "区分値(上位)", "品質所見"]
    for i, h in enumerate(ch, 1):
        _c(ws, r, i, h, font=F_H, fill=FILL_H, al=CC)
    ws.row_dimensions[r].height = 28
    hr = r
    for i, (cname, info) in enumerate(res["cols_out"].items(), 1):
        rr = hr + i
        st = FILL_STRIPE if i % 2 == 0 else None
        mn = md = mx = sd = ""
        if info["num"]:
            n = info["num"]
            mn, md, mx, sd = num(n["min"]), num(n["median"]), num(n["max"]), num(n["std"])
        elif info["date"]:
            mn, mx = info["date"]["min"], info["date"]["max"]
        vals = ""
        if info["values"]:
            vals = " / ".join(f"{v}({num(cnt)})" for v, cnt in info["values"][:6])
            if len(info["values"]) > 6:
                vals += " …"
        note = "▲全NULL(未投入)" if info["null_pct"] == 100 else ""
        cells = [i, cname, info["type"], f"{info['null_pct']:.0f}%",
                 f"{info['unique_pct']:.0f}%", num(info["dist"]), mn, md, mx, sd, vals, note]
        for j, v in enumerate(cells, 1):
            al = CC if j in (1, 3, 4, 5, 6, 7, 8, 9, 10) else LT
            _c(ws, rr, j, v, fill=st, al=al)
        ws.row_dimensions[rr].height = 20
    r = hr + len(res["cols_out"]) + 2

    # Phase3
    _c(ws, r, 1, "■ Phase3 セグメント別", font=F_SEC, border=False)
    r += 1
    _c(ws, r, 1, "区分値の少ない列を「軸」に、データを区分ごとに分けて件数構成を確認。"
       "さらに各列のNULL率(欠損割合)が区分間でどれだけ違うかを見る（差が大きい＝特定区分でのみ欠落の可能性）。",
       font=F_NOTE, border=False)
    r += 1
    kinds = {
        "null": ("NULL率(空欄割合)", "特定の区分でだけ値が無い（計算/取得されていない）可能性"),
        "unique": ("一意率(値の多様性)", "特定の区分でだけ値の種類が偏る/逆に多様な可能性"),
        "avg": ("平均値", "区分によって値の水準が大きく異なる"),
    }
    if res["segments"]:
        # 軸ごとの区分別件数（コンパクトに1行）
        for s in res["segments"]:
            top = " / ".join(f"{x['seg']}: {num(x['cnt'])}件({x['pct']:.0f}%)" for x in s["segs"][:8])
            _c(ws, r, 1, f"軸「{s['axis']}」の区分別件数: {top}", font=F_B, border=False)
            r += 1
        r += 1
        # 差が大きい所見をスコア順に集約（最大4件まで・各々を表で表示）
        findings = []
        for s in res["segments"]:
            for fl in s["flagged"]:
                score = (fl["max"] / fl["min"] - 1) * 50 if fl["kind"] == "avg" and fl["min"] else \
                    (fl["max"] - fl["min"])
                findings.append((score, s["axis"], fl))
        findings.sort(key=lambda x: -x[0])
        shown = findings[:4]
        if findings:
            extra = f"（差大 {len(findings)}件中 上位{len(shown)}件を表示）" if len(findings) > 4 else ""
            _c(ws, r, 1, f"区分間で差が大きい列{extra}:", font=F_B, border=False)
            r += 1
        for _score, axis, fl in shown:
            label, expl = kinds[fl["kind"]]
            rng = (f"{fl['min']:.0f}%〜{fl['max']:.0f}%" if fl["kind"] != "avg"
                   else f"{num(fl['min'])}〜{num(fl['max'])}")
            _c(ws, r, 1, f"▲ 列「{fl['col']}」の{label}（軸「{axis}」） … {rng} と差大：{expl}",
               font=F_B, border=False)
            r += 1
            # 小表: 区分 | 値
            _c(ws, r, 1, f"区分（{axis}）", font=F_H, fill=FILL_H, al=CC)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            _c(ws, r, 2, fill=FILL_H)
            _c(ws, r, 3, label, font=F_H, fill=FILL_H, al=CC)
            r += 1
            for j, (seg, v) in enumerate(fl["by_seg"]):
                disp = num(v) if fl["kind"] == "avg" else (f"{v:.0f}%" if v is not None else "")
                st = FILL_STRIPE if j % 2 == 1 else None
                _c(ws, r, 1, seg, fill=st, al=LC)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                _c(ws, r, 2, fill=st)
                _c(ws, r, 3, disp, fill=st, al=CC)
                r += 1
            r += 1
    else:
        _c(ws, r, 1, "（区分の軸にできる低カーディナリティ列がないため分析なし）", font=F_NOTE, border=False)
        r += 1
    r += 1

    # Phase4
    _c(ws, r, 1, "■ Phase4 品質チェック", font=F_SEC, border=False)
    r += 1
    _c(ws, r, 1, "データの異常・違和感を検出（各項目に説明付き）。", font=F_NOTE, border=False)
    r += 1
    for line in _quality_lines(res["quality"], res["cols_out"]):
        _c(ws, r, 1, line, font=F_B, border=False)
        r += 1
    r += 1

    # テーブル間整合
    _c(ws, r, 1, "■ テーブル間整合（このテーブル関連）", font=F_SEC, border=False)
    r += 1
    _c(ws, r, 1, "共有するキー列が、参照先（親テーブル＝そのキーが一意なマスタ等）と紐付くかを確認。"
       "「孤児」＝子テーブルにある値のうち、親テーブルに存在しない値の件数"
       "（表記ゆれ・未登録・名寄せ漏れで参照先と一致しない）。○整合＝孤児ゼロ。", font=F_NOTE, border=False)
    r += 1
    if cross_lines:
        for line in cross_lines:
            _c(ws, r, 1, line, font=F_B, border=False)
            r += 1
    else:
        _c(ws, r, 1, "（関連する共有キーは検出されませんでした）", font=F_NOTE, border=False)
        r += 1
    autofit(ws)


def build_workbook(results, cross, scope, project, gen_date, out_path,
                   logical_names=None):
    """results: profile()の戻り値リスト, cross: クロスチェック行のリスト（文字列）"""
    logical_names = logical_names or {}
    for r in results:
        r["logical"] = logical_names.get(r["table"], "")

    wb = Workbook()
    build_cover(wb, scope, project, gen_date, len(results))

    used = {"表紙", "テーブル一覧"}
    sheet_of = {}
    for r in sorted(results, key=sort_key):
        sheet_of[(r["dataset"], r["table"])] = sheet_name_for(r["table"], used)

    build_list(wb, results, sheet_of)

    # クロス行をテーブル単位に振り分け（行テキストに当該テーブル名を含むものを関連とみなす）
    for r in sorted(results, key=sort_key):
        rel = [ln for ln in cross if r["table"] in ln]
        build_detail(wb, r, sheet_of[(r["dataset"], r["table"])], rel)

    import os
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
